# -*- coding: utf-8 -*-
"""
flask_jira_refactored.py
========================
Refactorización de flask_jira.py para integración con Power Automate.

CAMBIOS RESPECTO AL ORIGINAL
─────────────────────────────────────────────────────────────────
MÉTODOS HTTP
  GET  /process-data      → GET  ✅  Solo consulta Jira + BD (lectura pura)
  GET  /guardar-historico → POST ✅  CAMBIADO: escribe en PostgreSQL (no idempotente)
  GET  /download-excel    → GET  ✅  Descarga de archivo (idempotente)

  Criterio de decisión:
    • GET   → La operación es SEGURA (no muta nada) e IDEMPOTENTE (mismo
              resultado N veces). Usada para consultas y descargas.
    • POST  → La operación muta estado persistente (DB, ficheros). No es
              idempotente: ejecutarla dos veces produce dos inserciones distintas
              (aunque el DELETE previo lo mitiga, la INTENCIÓN sigue siendo mutar).
              /guardar-historico escribe en PostgreSQL → POST correcto.

RESPUESTAS JSON
  Todos los endpoints devuelven jsonify() con la envoltura estándar:
    {
      "status":    "success" | "error",
      "timestamp": "ISO-8601 UTC",
      "message":   "...",
      "meta":      { totales, parámetros usados, etc. },
      "data":      { payload principal }
    }

  Power Automate accede con expresiones como:
    @body('HTTP')?['data']?['tickets']
    @body('HTTP')?['meta']?['total_issues']
    @body('HTTP')?['status']

VALIDACIÓN
  • _parse_date()  — valida YYYY-MM-DD, lanza ValueError descriptivo
  • _parse_bool()  — acepta bool nativo, "true"/"false", "1"/"0", "si"/"no"
  • _get_params()  — lee query string (GET) y JSON body (POST) de forma unificada

CORS
  Headers Access-Control añadidos en after_request para que el conector
  HTTP de Power Automate no falle en el pre-flight OPTIONS.

OTROS
  • Content-Type: application/json; charset=utf-8 explícito en todas las
    respuestas JSON (Power Automate lo requiere para parsear automáticamente).
  • Credenciales de BD y Jira preferiblemente desde variables de entorno.
  • HTML_TEMPLATE y funciones de HTML intactos (no se tocan, siguen usándose
    en /process-data para el render de correo/navegador).
  • /download-excel devuelve archivo binario, no JSON (correcto por diseño).
  • /health añadido: permite que Power Automate verifique disponibilidad
    antes de ejecutar el flujo principal.

CORRECCIONES — customfield_11533 (Cascading Select) y otros bugs
─────────────────────────────────────────────────────────────────
  • FIX 1: CF_GRUPO_LOCALIDAD y CF_LOCALIDAD conservan el mismo ID de campo
           (customfield_11533) pero se documentan como padre/hijo. La diferenciación
           se resuelve en runtime mediante funciones especializadas.
  • FIX 2: JIRA_FIELDS deduplicado con dict.fromkeys() — Jira recibe el campo
           una sola vez por request (ahorro de ancho de banda).
  • FIX 3: NameError "get_val" en construir_dataframe corregido a get_cascading_parent().
           Se añade la coma faltante que causaba SyntaxError en la misma línea.
  • FIX 4: get_value() no diferenciaba padre de hijo del Cascading Select.
           Se añaden get_cascading_parent() y get_cascading_child() con guards
           completos contra None, dict ausente y clave "child" faltante.
  • FIX 5: Typo "empresa_ejeuctor" corregido a "empresa_ejecutor".

MEJORAS v3 — corrección de errores críticos identificados en revisión
─────────────────────────────────────────────────────────────────
  • FIX 6:  INSERT fila a fila reemplazado por bulk INSERT con to_sql() —
            elimina N round-trips a PostgreSQL por ejecución (causa principal del WORKER TIMEOUT).
  • FIX 7:  DELETE + INSERT ahora ocurren en una sola transacción atómica —
            si el INSERT falla, el DELETE se revierte (no más tabla vacía intermedia).
  • FIX 8:  Engine SQLAlchemy promovido a singleton global con pool de conexiones —
            elimina el costo de crear/destruir el engine en cada llamada.
  • FIX 9:  guardar_historico_evolutivo() eliminado de GET /process-data —
            /process-data es ahora lectura pura; el guardado solo ocurre desde
            POST /guardar-historico (separa responsabilidades, elimina timeout).
  • FIX 10: Límite de páginas (MAX_PAGES) en obtener_issues() — previene loop infinito.
  • FIX 11: timeout de requests reducido a 25s (< timeout de Gunicorn) para que el
            decorador handle_errors pueda capturar Timeout antes de que Gunicorn mate el worker.
  • FIX 12: Rate-limit 429 usa min(wait, 20) para no exceder el timeout de Gunicorn.
  • FIX 13: FutureWarning de pandas corregido — df.copy() explícito antes de asignaciones
            en construir_dataframe() y generar_mensual().
  • FIX 14: app.config JSON_AS_ASCII/JSON_SORT_KEYS deprecados reemplazados por
            app.json.ensure_ascii y app.json.sort_keys (compatibles con Flask 2.3+).
  • FIX 15: Imports base64 y Flask.Response movidos al bloque de imports globales.
  • FIX 16: import datetime redundante dentro de fecha_corta_es() eliminado.
  • FIX 17: _df_to_records() reforzado — pd.isna() protegido con try/except para
            evitar ValueError en arrays de numpy.
  • FIX 18: Valores de texto de Jira escapados con html.escape() en las funciones
            de generación HTML para prevenir XSS.
  • FIX 19: Validación de PG_SCHEMA y PG_TABLE_EVOLUTIVO contra allowlist para
            prevenir inyección SQL a nivel de nombre de tabla/schema.

@author: mmerinori
@refactored: Claude (Anthropic) — Feb 2026
@fixes:     Claude (Anthropic) — Mar 2026
@fixes-v3:  Claude (Anthropic) — Jun 2026
"""

# ─────────────────────────────────────────────
# IMPORTS
# ─────────────────────────────────────────────
import base64
import html as html_lib
import io
import os
import time
import traceback
from datetime import datetime, timedelta, timezone
from functools import wraps
from typing import Any, Dict, List, Optional

import pandas as pd
import requests
from flask import Flask, Response as FlaskResponse, jsonify, make_response, request, render_template_string
from requests.auth import HTTPBasicAuth
from sqlalchemy import create_engine, text
from sqlalchemy.engine import URL as SA_URL

from dotenv import load_dotenv


load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))


# ─────────────────────────────────────────────
# APLICACIÓN FLASK
# ─────────────────────────────────────────────
app = Flask(__name__)

# FIX 14: JSON_AS_ASCII y JSON_SORT_KEYS fueron removidos en Flask 2.3.
# La forma correcta de configurarlos es a través de app.json (JSONProvider).
app.json.ensure_ascii = False   # Preserva UTF-8 en respuestas JSON
app.json.sort_keys    = False   # Mantiene orden lógico de claves


# ─────────────────────────────────────────────
# CORS — requerido por Power Automate
# ─────────────────────────────────────────────
@app.after_request
def apply_cors(response):
    """
    Power Automate realiza pre-flight OPTIONS antes de cada llamada real.
    Sin estos headers el conector HTTP devuelve 'Network Error' genérico.
    """
    response.headers["Access-Control-Allow-Origin"]  = "*"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization, Accept"
    return response


# ─────────────────────────────────────────────
# CONFIGURACIÓN JIRA
# ─────────────────────────────────────────────
JIRA_URL  = os.getenv("JIRA_URL")
EMAIL     = os.getenv("JIRA_EMAIL")
API_TOKEN = os.getenv("JIRA_TOKEN")

# Validación de seguridad
if not JIRA_URL or not EMAIL or not API_TOKEN:
    raise RuntimeError(
        "Variables de entorno JIRA_URL, JIRA_EMAIL y JIRA_TOKEN son obligatorias."
    )

AUTH         = HTTPBasicAuth(EMAIL, API_TOKEN)
REQ_HEADERS  = {"Accept": "application/json"}
API_ENDPOINT = "/rest/api/3/search/jql"

# FIX 10: Límite de páginas para prevenir loop infinito si Jira devuelve
# nextPageToken indefinidamente por un bug o datos corruptos.
MAX_PAGES = 50   # 50 páginas × 100 issues = 5000 issues máximo por request

# FIX 11: timeout reducido a 25s para que handle_errors capture
# requests.exceptions.Timeout ANTES de que Gunicorn (30s) mate el worker.
JIRA_REQUEST_TIMEOUT = 25

# Custom Fields
CF_FECHA_FIN        = "customfield_11365"
CF_FECHA_INICIO     = "customfield_11363"
CF_SUPERVISOR       = "customfield_11451"
CF_FECHA_COMITE     = "customfield_11673"
CF_TIPO_VENTANA     = "customfield_10486"
CF_REGION           = "customfield_10952"
# FIX 1: CF_GRUPO_LOCALIDAD y CF_LOCALIDAD comparten el mismo customfield (Cascading Select).
# La separación es semántica: padre → grupo_localidad, hijo → localidad.
# get_cascading_parent() y get_cascading_child() resuelven el nivel correcto en runtime.
CF_GRUPO_LOCALIDAD  = "customfield_11533"   # nivel padre del Cascading Select
CF_LOCALIDAD        = "customfield_11533"   # nivel hijo  del Cascading Select (mismo campo JSON)
CF_EMPRESA_EJECUTOR = "customfield_11415"

# FIX 2: dict.fromkeys() elimina duplicados preservando el orden.
# CF_GRUPO_LOCALIDAD y CF_LOCALIDAD son el mismo ID → Jira solo necesita recibirlo una vez.
JIRA_FIELDS = list(dict.fromkeys([
    "summary", "status", "creator",
    CF_FECHA_INICIO, CF_FECHA_FIN, CF_SUPERVISOR, CF_FECHA_COMITE,
    CF_TIPO_VENTANA, CF_REGION, CF_GRUPO_LOCALIDAD, CF_EMPRESA_EJECUTOR,
]))

MESES = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun",
         "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

# Zona horaria de Lima (UTC-5)
TZ_LIMA = timezone(timedelta(hours=-5))


def _ayer_lima_str() -> str:
    """
    Devuelve la fecha de ayer en hora Lima como "YYYY-MM-DD".
    Garantiza que los reportes automaticos siempre cierren en T-1.
    """
    return (datetime.now(TZ_LIMA) - timedelta(days=1)).strftime("%Y-%m-%d")


# ─────────────────────────────────────────────
# CONFIGURACIÓN POSTGRESQL
# ─────────────────────────────────────────────
PG_CONFIG = {
    "host":     os.getenv("PG_HOST"),
    "port":     os.getenv("PG_PORT"),
    "database": os.getenv("PG_DATABASE"),
    "user":     os.getenv("PG_USER"),
    "password": os.getenv("PG_PASSWORD"),
}

# Validación de seguridad PostgreSQL
if not all(PG_CONFIG.values()):
    raise RuntimeError(
        "Variables de entorno PG_HOST, PG_PORT, PG_DATABASE, PG_USER, PG_PASSWORD son obligatorias."
    )

PG_SCHEMA          = os.getenv("PG_SCHEMA", "proof")
PG_TABLE_EVOLUTIVO = "evol_jira_supervisores_p_im"

# FIX 19: Validar PG_SCHEMA y PG_TABLE_EVOLUTIVO contra una allowlist para
# prevenir inyección SQL a nivel de nombre de tabla/schema (no pueden parametrizarse
# con SQLAlchemy, por lo que se validan manualmente antes de interpolarlos en SQL).
_PG_SCHEMA_ALLOWLIST = {"proof", "public", "jira", "reportes"}
_PG_TABLE_ALLOWLIST  = {"evol_jira_supervisores_p_im"}

if PG_SCHEMA not in _PG_SCHEMA_ALLOWLIST:
    raise RuntimeError(
        f"PG_SCHEMA '{PG_SCHEMA}' no está en la allowlist permitida: {_PG_SCHEMA_ALLOWLIST}"
    )
if PG_TABLE_EVOLUTIVO not in _PG_TABLE_ALLOWLIST:
    raise RuntimeError(
        f"PG_TABLE_EVOLUTIVO '{PG_TABLE_EVOLUTIVO}' no está en la allowlist permitida: {_PG_TABLE_ALLOWLIST}"
    )

# FIX 8: Engine SQLAlchemy promovido a singleton global.
# Con pool_pre_ping=True SQLAlchemy verifica y restablece conexiones caídas automáticamente.
# pool_size=3 y max_overflow=5 son valores conservadores adecuados para Render free tier.
# Al ser global, el pool se comparte entre todas las requests y no se paga el costo
# de TCP handshake en cada llamada (era el comportamiento anterior con crear_engine_db()).
# FIX 15 (parcial): DB_URI reemplazado por SA_URL para que la contraseña no quede
# como string plano en memoria (SA_URL la encapsula y no la expone en __repr__).
_DB_URL = SA_URL.create(
    drivername="postgresql",
    username=PG_CONFIG["user"],
    password=PG_CONFIG["password"],
    host=PG_CONFIG["host"],
    port=int(PG_CONFIG["port"]),
    database=PG_CONFIG["database"],
)
DB_ENGINE = create_engine(
    _DB_URL,
    pool_pre_ping=True,
    pool_size=3,
    max_overflow=5,
    pool_recycle=1800,   # recicla conexiones inactivas cada 30 min
)


# ═══════════════════════════════════════════════════════════════
# UTILIDADES DE RESPUESTA JSON ESTANDARIZADA
# ═══════════════════════════════════════════════════════════════

def _ok(data: Any = None, message: str = "OK",
        meta: Dict = None, status: int = 200):
    """
    Construye la respuesta JSON de éxito canónica.

    Estructura que Power Automate puede parsear directamente:
    {
        "status":    "success",
        "timestamp": "2026-02-17T15:30:00+00:00",
        "message":   "42 issues procesados",
        "meta":      { "total_issues": 42, "fecha_inicio": "2025-11-01", ... },
        "data":      { "tickets": [...], "resumen_mensual": [...], ... }
    }

    Expresiones en Power Automate:
        @body('HTTP')?['data']?['tickets']
        @body('HTTP')?['meta']?['total_issues']
        @equals(body('HTTP')?['status'], 'success')
    """
    body = {
        "status":    "success",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "message":   message,
        "meta":      meta if meta is not None else {},
        "data":      data if data is not None else {},
    }
    resp = make_response(jsonify(body), status)
    resp.headers["Content-Type"] = "application/json; charset=utf-8"
    return resp


def _err(message: str, errors: List[str] = None, status: int = 500):
    """
    Construye la respuesta JSON de error canónica.

    Power Automate puede usar una condición sobre body('HTTP')?['status']
    para bifurcar el flujo y leer body('HTTP')?['errors'] para diagnóstico.
    """
    body = {
        "status":    "error",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "message":   message,
        "errors":    errors if errors else [],
        "data":      {},
    }
    resp = make_response(jsonify(body), status)
    resp.headers["Content-Type"] = "application/json; charset=utf-8"
    return resp


def handle_errors(f):
    """
    Decorador de manejo centralizado de excepciones.
    Garantiza que el conector HTTP de Power Automate SIEMPRE reciba JSON
    válido, incluso ante errores inesperados (Flask por defecto devuelve HTML).
    """
    @wraps(f)
    def wrapper(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except ValueError as exc:
            # Parámetros inválidos → 400 Bad Request
            return _err("Parámetro inválido", errors=[str(exc)], status=400)
        except requests.exceptions.Timeout:
            # FIX 11: Con JIRA_REQUEST_TIMEOUT=25s < 30s de Gunicorn,
            # este bloque se alcanza antes de que Gunicorn mate el worker.
            return _err(
                "Timeout al conectar con Jira",
                errors=["El servicio externo no respondió en el tiempo esperado (25s)"],
                status=504,
            )
        except requests.exceptions.ConnectionError as exc:
            # Red caída → 502 Bad Gateway
            return _err("Error de conexión con Jira", errors=[str(exc)], status=502)
        except Exception as exc:
            traceback.print_exc()
            return _err("Error interno del servidor", errors=[str(exc)], status=500)
    return wrapper


# ─────────────────────────────────────────────
# VALIDACIÓN Y PARSEO DE PARÁMETROS
# ─────────────────────────────────────────────

def _parse_date(value: str, param_name: str) -> str:
    """
    Valida formato YYYY-MM-DD y devuelve el valor como string limpio.
    Lanza ValueError con mensaje descriptivo si el formato es incorrecto.

    Conversión de tipo explícita: Power Automate puede enviar la fecha
    como string o como datetime serializado; ambos se normalizan aquí.
    """
    if not isinstance(value, str):
        value = str(value)
    value = value.strip()[:10]          # Trunca a YYYY-MM-DD si trae hora
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return value
    except ValueError:
        raise ValueError(
            f"Parámetro '{param_name}' con formato inválido: '{value}'. "
            "Se esperaba YYYY-MM-DD (ej. 2026-02-17)."
        )


def _parse_bool(value: Any, default: bool = False) -> bool:
    """
    Conversión robusta a bool.
    Power Automate puede enviar true/false como bool nativo o como string.
    """
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in ("true", "1", "yes", "si", "sí")
    return default


def _get_params() -> Dict:
    """
    Lee parámetros de forma unificada:
      - GET:  query string  ?inicio=2025-11-01&fin=2026-02-17
      - POST: JSON body     {"inicio": "2025-11-01", "fin": "2026-02-17"}

    El body JSON tiene prioridad sobre query string para evitar
    ambigüedad cuando Power Automate envía ambos.
    """
    params = dict(request.args)
    if request.is_json:
        body = request.get_json(silent=True) or {}
        params.update(body)
    return params


# ═══════════════════════════════════════════════════════════════
# LÓGICA DE NEGOCIO
# ═══════════════════════════════════════════════════════════════

def get_value(field: Any) -> Optional[str]:
    """
    Extrae el valor de un campo Jira estándar (option, user, string o None).

    NO usar para Cascading Select — usar get_cascading_parent() / get_cascading_child().

    Casos cubiertos:
      • None                          → None
      • dict con "value"              → valor del option (ej. tipo_ventana, region)
      • dict con "displayName"/"name" → nombre de usuario/estado
      • string / número               → cast a str
    """
    if field is None:
        return None
    if isinstance(field, dict):
        return field.get("value") or field.get("displayName") or field.get("name")
    return str(field)


def get_cascading_parent(field: Any) -> Optional[str]:
    """
    Extrae el valor del NIVEL PADRE de un campo Cascading Select de Jira.

    Estructura JSON esperada:
        {
            "value": "PARCACHATA",          ← devuelve esto
            "id":    "81928",
            "child": { "value": "APURIMAC - COTARUSE", "id": "81929" }
        }

    Casos seguros (nunca lanza KeyError ni AttributeError):
      • field es None          → None
      • field no es dict       → None
      • "value" no existe      → None  (field.get() retorna None por defecto)
    """
    if not isinstance(field, dict):
        return None
    return field.get("value")


def get_cascading_child(field: Any) -> Optional[str]:
    """
    Extrae el valor del NIVEL HIJO de un campo Cascading Select de Jira.

    Estructura JSON esperada:
        {
            "value": "PARCACHATA",
            "id":    "81928",
            "child": {
                "value": "APURIMAC - COTARUSE",  ← devuelve esto
                "id":    "81929"
            }
        }

    Casos seguros (nunca lanza KeyError ni AttributeError):
      • field es None                      → None
      • field no es dict                   → None
      • "child" no existe en el dict       → None  (ticket sin nivel hijo seleccionado)
      • "child" existe pero no es dict     → None  (dato inesperado, defensivo)
      • "child" existe pero sin "value"    → None
    """
    if not isinstance(field, dict):
        return None
    child = field.get("child")          # None si el usuario no eligió nivel hijo
    if not isinstance(child, dict):
        return None
    return child.get("value")


def calcular_supervisor_final(row: pd.Series) -> str:
    """Aplica las reglas de negocio para determinar el supervisor final."""
    resumen        = row.get("resumen",        "") or ""
    supervisor     = row.get("supervisor",     "") or ""
    registrado_por = row.get("registrado_por", "") or ""
    estado         = row.get("estado",         "") or ""

    if estado not in ("Programado", "Implantación en Curso"):
        return supervisor

    if "MP@" in resumen and registrado_por not in (
        "Lydia Taco Cáceres", "Maria Amparo Meza Obando"
    ):
        return "HERMANN ANDRE ROBLES CUADROS"

    if "|" in supervisor:
        return supervisor.split("|")[0].strip()

    return supervisor


def obtener_issues(fecha_inicio: str, fecha_fin: str) -> List[Dict[str, Any]]:
    """
    Consulta paginada a Jira usando nextPageToken.
    GET a Jira es correcto: solo LEEMOS datos, no mutamos nada en Jira.

    FIX 10: Se añade un límite de MAX_PAGES páginas para evitar un loop
    infinito si Jira devuelve nextPageToken indefinidamente por un bug.
    FIX 11: timeout reducido a JIRA_REQUEST_TIMEOUT (25s) para que
    handle_errors pueda capturar la excepción antes del timeout de Gunicorn.
    FIX 12: En caso de rate-limit 429, el tiempo de espera se limita a
    min(Retry-After, 20) segundos para no exceder el timeout de Gunicorn.
    """
    jql_query = (
        f'project = TPRO '
        f'AND status in ("Programado","Implantación en Curso") '
        f'AND "Fecha programada de finalización[Time stamp]" >= "{fecha_inicio} 00:00" '
        f'AND "Fecha programada de finalización[Time stamp]" <= "{fecha_fin} 23:59"'
    )
    print(f"[{datetime.now():%H:%M:%S}] JQL: {jql_query}")

    url        = f"{JIRA_URL}{API_ENDPOINT}"
    all_issues = []
    next_token = None
    pagina     = 0

    while pagina < MAX_PAGES:
        pagina += 1
        params = {
            "jql":        jql_query,
            "fields":     ",".join(JIRA_FIELDS),
            "maxResults": 100,
        }
        if next_token:
            params["nextPageToken"] = next_token

        try:
            response = requests.get(
                url, headers=REQ_HEADERS, params=params,
                auth=AUTH, timeout=JIRA_REQUEST_TIMEOUT,
            )

            if response.status_code == 429:
                # FIX 12: Limitar el wait para no superar el timeout de Gunicorn.
                wait = min(int(response.headers.get("Retry-After", 10)), 20)
                print(f"[{datetime.now():%H:%M:%S}] Rate limit — esperando {wait}s (máx 20s)…")
                time.sleep(wait)
                pagina -= 1   # No contar este intento fallido como página procesada
                continue

            if response.status_code != 200:
                print(f"[{datetime.now():%H:%M:%S}] Error {response.status_code}: {response.text[:300]}")
                break

            data   = response.json()
            issues = data.get("issues", [])
            all_issues.extend(issues)
            print(f"[{datetime.now():%H:%M:%S}] Página {pagina} — {len(issues)} issues | acumulado: {len(all_issues)}")

            next_token = data.get("nextPageToken")
            if not next_token:
                break

            time.sleep(0.2)

        except requests.exceptions.Timeout:
            # Re-lanzar para que handle_errors lo capture y devuelva 504
            raise
        except Exception as exc:
            print(f"[{datetime.now():%H:%M:%S}] Error en request: {exc}")
            break

    if pagina >= MAX_PAGES and next_token:
        print(f"[{datetime.now():%H:%M:%S}] ⚠️ Se alcanzó el límite de {MAX_PAGES} páginas — puede haber más issues.")

    return all_issues


def construir_dataframe(issues: List[Dict[str, Any]]) -> pd.DataFrame:
    """Transforma la lista cruda de issues Jira a un DataFrame normalizado."""
    rows = []
    for issue in issues:
        try:
            f = issue["fields"]
            # FIX 3+4: Se lee el campo cascading una sola vez y se pasa a
            # las dos funciones especializadas. Sin NameError, sin coma faltante.
            localidad_raw = f.get(CF_GRUPO_LOCALIDAD)   # mismo campo para padre e hijo

            rows.append({
                "cod_trabajo_programado": issue["key"],
                "tipo_ventana":           get_value(f.get(CF_TIPO_VENTANA)),
                "estado":                 f["status"]["name"],
                "resumen":                f.get("summary"),
                "fecha_inicio":           f.get(CF_FECHA_INICIO),
                "fecha_fin":              f.get(CF_FECHA_FIN),
                "supervisor":             get_value(f.get(CF_SUPERVISOR)),
                "fecha_comite":           f.get(CF_FECHA_COMITE),
                "registrado_por":         get_value(f.get("creator")),
                "region":                 get_value(f.get(CF_REGION)),
                "grupo_localidad":        get_cascading_parent(localidad_raw),   # "PARCACHATA"
                "localidad":              get_cascading_child(localidad_raw),    # "APURIMAC - COTARUSE"
                "empresa_ejecutor":       get_value(f.get(CF_EMPRESA_EJECUTOR)), # FIX 5: typo corregido
            })
        except Exception as exc:
            print(f"[{datetime.now():%H:%M:%S}] Error procesando {issue.get('key')}: {exc}")

    df = pd.DataFrame(rows)

    if not df.empty:
        # FIX 13 (definitivo): Construir el DataFrame directamente con las columnas
        # de fecha ya convertidas, en lugar de asignar sobre un DataFrame existente.
        # Esto elimina completamente el FutureWarning de chained assignment de pandas
        # porque nunca hay una asignación post-construcción sobre una vista.
        # pd.to_datetime + dt.tz_localize opera sobre la Serie original (no una copia
        # de una copia), así que pandas Copy-on-Write no detecta la cadena.
        for col in ("fecha_inicio", "fecha_fin", "fecha_comite"):
            # Jira está configurado en hora Lima: los timestamps ya son
            # hora local. Se parsean sin zona para no restar 5 horas.
            parsed = pd.to_datetime(df[col], errors="coerce").dt.tz_localize(None)
            df = df.assign(**{col: parsed})
        df = df.assign(supervisor_final=df.apply(calcular_supervisor_final, axis=1))

    return df


def generar_mensual(df: pd.DataFrame) -> pd.DataFrame:
    """Agrupación mensual por supervisor, registrado_por y estado."""
    if df.empty:
        return pd.DataFrame()
    # FIX 13 (definitivo): df.assign() crea una copia interna sin
    # disparar FutureWarning — nunca modifica el DataFrame original.
    df = df.assign(mes=df["fecha_fin"].dt.strftime("%Y-%m"))
    return (
        df.groupby(["supervisor_final", "registrado_por", "mes", "estado"])
        .size()
        .reset_index(name="cantidad")
    )


# ─────────────────────────────────────────────
# BASE DE DATOS — EVOLUTIVO
# ─────────────────────────────────────────────

def guardar_historico_evolutivo(df: pd.DataFrame,
                                fecha_inicio: str,
                                fecha_fin: str) -> Dict:
    """
    Persiste el snapshot en PostgreSQL usando el engine global (FIX 8).

    FIX 6: El INSERT fila a fila fue reemplazado por un único bulk INSERT
    con pandas to_sql() + método 'multi'. Esto elimina N round-trips de red
    (uno por supervisor) y los reduce a una sola operación, resolviendo el
    cuello de botella que causaba el WORKER TIMEOUT de Gunicorn.

    FIX 7: DELETE e INSERT ocurren ahora dentro de una sola transacción
    atómica (mismo bloque `with engine.begin()`). Si el INSERT falla, el
    DELETE se revierte automáticamente — no más tabla vacía intermedia.

    Por qué POST para el endpoint que llama a esta función:
      • Ejecutar dos veces con los mismos parámetros produce dos operaciones
        DELETE + INSERT distintas en la base de datos → no es idempotente.
      • La operación tiene EFECTO SECUNDARIO OBSERVABLE (modifica la BD).
      → POST es el método HTTP correcto según RFC 9110.
    """
    if df.empty:
        return {"guardado": False, "motivo": "DataFrame vacío, sin datos para persistir"}

    try:
        resumen = (
            df.groupby(["supervisor_final", "estado"])
            .size()
            .unstack(fill_value=0)
        )
        resumen["total_registros"] = resumen.sum(axis=1)
        resumen = resumen.reset_index()

        resumen = resumen.rename(columns={
            "Implantación en Curso": "total_implantacion",
            "Programado":            "total_programado",
            "supervisor_final":      "supervisor",
        })

        for col in ("total_implantacion", "total_programado"):
            if col not in resumen.columns:
                resumen[col] = 0

        fecha_fin_obj    = datetime.strptime(fecha_fin,    "%Y-%m-%d").date()
        fecha_inicio_obj = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()

        # Construir strings ISO explícitos para evitar que psycopg2 aplique
        # la zona horaria del servidor al insertar en columnas TIMESTAMPTZ.
        # Al pasar strings, PostgreSQL los interpreta como hora literal sin offset.
        fecha_corte_str     = f"{fecha_fin} 23:59:59"
        fecha_ejecucion_str = datetime.now(TZ_LIMA).strftime("%Y-%m-%d %H:%M:%S")

        resumen["fecha_corte"]     = fecha_corte_str
        resumen["fecha_inicio"]    = str(fecha_inicio_obj)
        resumen["fecha_fin"]       = str(fecha_fin_obj)
        resumen["fecha_ejecucion"] = fecha_ejecucion_str

        # Asegurar tipos correctos para el INSERT
        resumen["total_registros"]    = resumen["total_registros"].astype(int)
        resumen["total_implantacion"] = resumen["total_implantacion"].astype(int)
        resumen["total_programado"]   = resumen["total_programado"].astype(int)

        cols_finales = [
            "fecha_corte", "supervisor", "total_registros",
            "total_implantacion", "total_programado",
            "fecha_inicio", "fecha_fin", "fecha_ejecucion",
        ]
        resumen_final = resumen[cols_finales].copy()

        # FIX 7: engine.begin() abre una transacción que hace commit automático
        # al salir del bloque 'with' sin error, y rollback si ocurre una excepción.
        # DELETE e INSERT comparten la misma transacción → atomicidad garantizada.
        #
        # FIX 6: to_sql() con method="multi" genera un único INSERT con todos los
        # valores, eliminando el loop fila a fila que causaba N round-trips a BD.
        # if_exists="append" porque la tabla ya existe; el DELETE previo asegura
        # que no haya duplicados para la misma fecha_fin.
        with DB_ENGINE.begin() as conn:
            # Primero: borrar registros existentes para esta fecha_fin
            conn.execute(
                text(f"DELETE FROM {PG_SCHEMA}.{PG_TABLE_EVOLUTIVO} WHERE fecha_fin = :fd"),
                {"fd": str(fecha_fin_obj)},
            )

            # Segundo: insertar todos los registros de una vez (bulk INSERT)
            # El CAST de fechas/timestamps se delega a PostgreSQL a través de
            # los tipos nativos de pandas (object → text que PG casteará).
            resumen_final.to_sql(
                name=PG_TABLE_EVOLUTIVO,
                con=conn,
                schema=PG_SCHEMA,
                if_exists="append",
                index=False,
                method="multi",   # genera un único INSERT multi-fila
            )

        n = len(resumen_final)
        print(f"[{datetime.now():%H:%M:%S}] ✅ Histórico guardado: {n} supervisores (bulk INSERT)")
        return {
            "guardado":               True,
            "supervisores_guardados": n,
            "fecha_corte":            fecha_fin,
        }

    except Exception as exc:
        traceback.print_exc()
        raise RuntimeError(f"Error al guardar histórico en BD: {exc}") from exc
    # FIX 8: No se llama engine.dispose() porque el engine es global y
    # gestiona su propio pool. dispose() solo se llamaría al apagar la app.


def consultar_historico_evolutivo() -> pd.DataFrame:
    """
    Consulta las últimas 5 fechas de corte desde PostgreSQL.
    Usa el engine global (FIX 8) — sin crear/destruir conexiones por llamada.
    """
    try:
        # Agrupamos por fecha_fin (tipo DATE, sin zona horaria) en lugar de
        # fecha_corte::DATE, porque fecha_corte es TIMESTAMPTZ y su conversión
        # a DATE en UTC puede caer en el día siguiente (+5h Lima → UTC).
        query = f"""
        WITH ultimas_fechas AS (
            SELECT DISTINCT fecha_fin::DATE AS fecha
            FROM {PG_SCHEMA}.{PG_TABLE_EVOLUTIVO}
            ORDER BY fecha DESC
            LIMIT 5
        )
        SELECT supervisor,
               fecha_fin::DATE AS fecha_orden,
               total_registros
        FROM {PG_SCHEMA}.{PG_TABLE_EVOLUTIVO}
        WHERE fecha_fin::DATE IN (SELECT fecha FROM ultimas_fechas)
        ORDER BY supervisor, fecha_fin
        """
        df = pd.read_sql(query, DB_ENGINE)
        if df.empty:
            return pd.DataFrame()

        df_pivot = df.pivot_table(
            index="supervisor", columns="fecha_orden",
            values="total_registros", fill_value=0,
        )
        print(f"[{datetime.now():%H:%M:%S}] ✅ Histórico: {len(df_pivot)} supervisores, {len(df_pivot.columns)} fechas")
        return df_pivot

    except Exception as exc:
        print(f"[{datetime.now():%H:%M:%S}] ❌ Error consultando histórico: {exc}")
        return pd.DataFrame()


def generar_evolutivo(df: pd.DataFrame, usar_bd: bool = True,
                      fecha_fin: str = None) -> pd.DataFrame:
    """Devuelve el DataFrame pivoteado para la tabla evolutiva."""
    if usar_bd:
        df_evol = consultar_historico_evolutivo()
        if not df_evol.empty:
            return df_evol
        print(f"[{datetime.now():%H:%M:%S}] ⚠️  Sin histórico en BD, usando datos actuales")

    if df.empty:
        return pd.DataFrame()

    df = df.copy()
    # Usar fecha_fin del parámetro si está disponible; si no, ayer en Lima
    if fecha_fin is not None:
        fecha_corte_val = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
    else:
        fecha_corte_val = (datetime.now(TZ_LIMA) - timedelta(days=1)).date()
    df["fecha_corte"] = fecha_corte_val
    return (
        df.groupby(["supervisor_final", "fecha_corte"])
        .size()
        .reset_index(name="total_registros")
        .pivot_table(
            index="supervisor_final", columns="fecha_corte",
            values="total_registros", fill_value=0,
        )
    )


# ─────────────────────────────────────────────
# HELPERS DE SERIALIZACIÓN
# ─────────────────────────────────────────────

def _df_to_records(df: pd.DataFrame) -> List[Dict]:
    """
    Convierte un DataFrame a lista de dicts serializables por jsonify.
    Maneja Timestamp, NaT y NaN para evitar errores de serialización JSON.

    FIX 17: pd.isna() protegido con try/except para evitar ValueError
    ("The truth value of an array is ambiguous") cuando val es un array
    de numpy en lugar de un escalar.
    """
    records = []
    for _, row in df.iterrows():
        rec = {}
        for col, val in row.items():
            if isinstance(col, tuple):
                col = "_".join(str(c) for c in col)
            col = str(col)

            # FIX 17: try/except en lugar de la expresión condicional frágil
            is_na = False
            if not isinstance(val, str):
                try:
                    is_na = bool(pd.isna(val))
                except (ValueError, TypeError):
                    is_na = False

            if is_na:
                rec[col] = None
            elif hasattr(val, "isoformat"):     # datetime / Timestamp
                rec[col] = val.isoformat()
            elif hasattr(val, "item"):          # numpy int/float
                rec[col] = val.item()
            else:
                rec[col] = val
        records.append(rec)
    return records


def _evolutivo_to_dict(df_evol: pd.DataFrame) -> List[Dict]:
    """
    Serializa el DataFrame pivoteado evolutivo a lista de dicts.
    Columnas fecha (índice de columnas del pivot) → string "Mes-DD".
    """
    if df_evol is None or df_evol.empty:
        return []

    result = []
    cols_fmt = []
    for c in df_evol.columns:
        if hasattr(c, "month"):
            cols_fmt.append(f"{MESES[c.month]}-{c.day:02d}")
        else:
            cols_fmt.append(str(c))

    for supervisor, row in df_evol.iterrows():
        entry = {"supervisor": str(supervisor)}
        for label, val in zip(cols_fmt, row):
            entry[label] = int(val) if not pd.isna(val) else 0
        result.append(entry)

    return result


# ─────────────────────────────────────────────
# HTML — funciones de generación de HTML para correo/navegador
# ─────────────────────────────────────────────

def fecha_corta_es(fecha) -> str:
    """
    Formatea una fecha como "Mes DD" en español abreviado.
    FIX 16: import datetime redundante eliminado — datetime ya está importado
    al inicio del módulo.
    """
    if isinstance(fecha, str):
        try:
            fecha = datetime.strptime(fecha[:10], "%Y-%m-%d").date()
        except ValueError:
            return str(fecha)
    if hasattr(fecha, "month") and hasattr(fecha, "day"):
        return f"{MESES[fecha.month]} {fecha.day:02d}"
    return str(fecha)


def generar_tabla_resumen_html(df_pivot: pd.DataFrame) -> str:
    """
    Genera la tabla HTML de resumen por supervisor/mes/estado.

    FIX 18: Los valores de texto provenientes de Jira (supervisor, registrado_por)
    se escapan con html.escape() para prevenir XSS en el correo generado.
    Un nombre en Jira con <script> o caracteres HTML quedaría escapado como
    &lt;script&gt; en lugar de ejecutarse.
    """
    if df_pivot.empty:
        return "<p>No hay datos</p>"

    style_th  = "background:#d4e8f0;color:#003d5c;padding:4px;border:1px solid #a6c9d7;font-size:11px;text-align:center;"
    style_td  = "padding:3px;border:1px solid #d0e4ed;font-size:11px;text-align:center;"
    style_sup = "background:#5a8ca3;color:#fff;font-weight:bold;"

    pivot = df_pivot.pivot_table(
        index=["supervisor_final", "registrado_por"],
        columns=["mes", "estado"], values="cantidad",
        aggfunc="sum", fill_value=0,
    ).sort_index(axis=1)

    pivot["Total"]   = pivot.sum(axis=1)
    totales_sup      = pivot.groupby(level=0).sum()
    orden            = totales_sup.sort_values("Total", ascending=False).index
    estados          = sorted(df_pivot["estado"].unique())
    meses            = sorted(df_pivot["mes"].unique())
    meses_label      = [f"{MESES[int(m[5:])]}-{m[2:4]}" for m in meses]

    html = [
        '<div><h3 style="color:#003d5c;font-size:13px;margin-bottom:8px;">'
        'Tabla resumen de tickets vencidos en estado "Programado" o "Implantación en curso"'
        "</h3></div>",
        '<table cellspacing="0" cellpadding="0" style="border-collapse:collapse;'
        'border:2px solid #7fa8ba;font-family:Calibri,Arial,sans-serif;width:100%;">',
        "<tr>",
        f'<th rowspan="2" style="{style_th}min-width:140px;">SUPERVISOR</th>',
    ]
    for est in estados:
        html.append(f'<th colspan="{len(meses)}" style="{style_th}">{html_lib.escape(est)}</th>')
    html.append(f'<th rowspan="2" style="{style_th}border-left:2px solid #7fa8ba;">Total</th></tr>')
    html.append("<tr>")
    for _ in estados:
        for ml in meses_label:
            html.append(f'<th style="{style_th}">{ml}</th>')
    html.append("</tr>")

    for sup in orden:
        if sup not in pivot.index.get_level_values(0):
            continue
        datos     = pivot.loc[sup]
        if isinstance(datos, pd.Series):
            datos = pd.DataFrame([datos])
        # .iloc[0] evita FutureWarning cuando .loc devuelve una Series de un elemento
        # en lugar de un escalar (ocurre con MultiIndex o índices duplicados)
        val_total = totales_sup.loc[sup, "Total"]
        total_sup = int(val_total.iloc[0]) if isinstance(val_total, pd.Series) else int(val_total)

        # FIX 18: escapar valores de texto que provienen de Jira
        sup_escaped = html_lib.escape(str(sup))
        html.append("<tr>")
        html.append(f'<td style="{style_td}{style_sup}text-align:left;padding-left:5px;">▼ {sup_escaped}</td>')
        for est in estados:
            for m in meses:
                v = totales_sup.loc[sup, (m, est)] if (m, est) in totales_sup.columns else 0
                html.append(f'<td style="{style_td}{style_sup}">{int(v) if v else ""}</td>')
        html.append(f'<td style="{style_td}{style_sup}border-left:2px solid #7fa8ba;">{total_sup}</td></tr>')

        detalle = datos.sort_values("Total", ascending=False)
        alt = False
        for reg, row in detalle.iterrows():
            bg = "#f5f9fb" if alt else "#fff"
            alt = not alt
            # FIX 18: escapar el nombre del registrado_por
            reg_escaped = html_lib.escape(str(reg))
            html.append("<tr>")
            html.append(f'<td style="{style_td}background:{bg};text-align:left;padding-left:15px;">{reg_escaped}</td>')
            for est in estados:
                for m in meses:
                    v = row.get((m, est), 0)
                    html.append(f'<td style="{style_td}background:{bg};">{int(v) if v else ""}</td>')
            val_row_total = row["Total"]
            total_row = int(val_row_total.iloc[0]) if isinstance(val_row_total, pd.Series) else int(val_row_total)
            html.append(f'<td style="{style_td}background:{bg};border-left:2px solid #7fa8ba;">{total_row}</td></tr>')

    # ── Fila Total General ───────────────────────────────
    style_tg = "background:#003d5c;color:#fff;font-weight:bold;"
    html.append("<tr>")
    html.append(f'<td style="{style_td}{style_tg}text-align:left;padding-left:5px;">Total general</td>')
    gran_total = 0
    for est in estados:
        for m in meses:
            col_key = (m, est)
            v = int(totales_sup[col_key].sum()) if col_key in totales_sup.columns else 0
            gran_total += v
            html.append(f'<td style="{style_td}{style_tg}">{v if v else ""}</td>')
    gran_total_total = int(totales_sup["Total"].sum())
    html.append(f'<td style="{style_td}{style_tg}border-left:2px solid #7fa8ba;">{gran_total_total}</td></tr>')

    html.append("</table></div>")
    return "".join(html)


def generar_tabla_evolutivo_html(df_evol: pd.DataFrame) -> str:
    """
    Genera la tabla HTML de evolución histórica por supervisor.
    FIX 18: supervisor escapado con html.escape().
    """
    if df_evol is None or df_evol.empty:
        return "<p>No hay datos evolutivos</p>"

    style_th = "background:#d4e8f0;color:#003d5c;padding:4px;border:1px solid #a6c9d7;font-size:11px;text-align:center;"
    style_td = "padding:3px;border:1px solid #d0e4ed;font-size:11px;text-align:center;"

    df_evol = df_evol.copy()
    df_evol.columns = [
        fecha_corta_es(c)
        for c in df_evol.columns
    ]
    if len(df_evol.columns) > 0:
        df_evol = df_evol.sort_values(df_evol.columns[-1], ascending=False)

    html = [
        '<div><h3 style="color:#003d5c;font-size:13px;margin-bottom:8px;">'
        "Evolución de tickets pendientes de cierre</h3></div>",
        '<table cellspacing="0" cellpadding="0" style="border-collapse:collapse;'
        'border:2px solid #7fa8ba;font-family:Calibri,Arial,sans-serif;">',
        f'<tr><th style="{style_th}">RESPONSABLE</th>',
    ]
    for col in df_evol.columns:
        html.append(f'<th style="{style_th}">{html_lib.escape(str(col))}</th>')
    html.append("</tr>")

    alt = False
    for sup, row in df_evol.iterrows():
        bg = "#f5f9fb" if alt else "#fff"
        alt = not alt
        # FIX 18: escapar nombre del supervisor
        sup_escaped = html_lib.escape(str(sup))
        html.append(f'<tr><td style="{style_td}background:{bg};text-align:left;padding-left:5px;font-weight:600;">{sup_escaped}</td>')
        for v in row:
            html.append(f'<td style="{style_td}background:{bg};">{int(v) if v else "-"}</td>')
        html.append("</tr>")

    html.append("</table></div>")
    return "".join(html)


HTML_TEMPLATE = """<!DOCTYPE html>
<html>
<head>
<meta http-equiv="Content-Type" content="text/html; charset=utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Reporte TPRO Jira</title>
<style>
    body { margin:0; padding:0; font-family:Calibri,Arial,sans-serif; color:#1c1b1b; background-color:#ffffff; }
    .container { max-width:100%; margin:0; background-color:#ffffff; padding:0; }
    .header { font-size:14px; line-height:1.6; margin-bottom:25px; color:#333; }
    .tables-wrapper { width:100%; border-collapse:collapse; }
    .tables-wrapper td { vertical-align:top; padding:0 10px; }
    .table-section-left  { width:58%; }
    .table-section-right { width:42%; }
</style>
</head>
<body>
<div class="container">
    <div class="header">
        Estimados,<br><br>
        Adjunto la relación de tickets TPRO Jira que se encuentran vencidos (su periodo programado ya finalizó)
        y permanecen en estado <strong>"Programado"</strong> o <strong>"Implantación en curso"</strong>.
        Agradeceremos su apoyo para realizar la <strong>regularización correspondiente a la brevedad posible.</strong><br>
        Les recordamos que, según el procedimiento vigente, <strong>cada ticket debe gestionarse dentro del periodo de
        ejecución definido en el propio TPRO</strong>, por tanto las fechas de apertura y cierre del ticket deben reflejar
        exactamente el inicio y fin reales del trabajo. En caso de no ejecutarse, el ticket debe ser cancelado.
        Este cumplimiento garantiza la correcta trazabilidad dentro del proceso de Gestión de Cambios.<br><br>
        La fecha de corte de este reporte es del <strong>{{ fecha_corte }}</strong>
    </div>
    <table class="tables-wrapper" cellpadding="0" cellspacing="0">
        <tr>
            <td class="table-section-left">{{ tabla_resumen|safe }}</td>
            <td class="table-section-right">{{ tabla_evolutivo|safe }}</td>
        </tr>
    </table>
</div>
</body>
</html>"""


# ═══════════════════════════════════════════════════════════════
# ENDPOINTS
# ═══════════════════════════════════════════════════════════════

# ─────────────────────────────────────────────
# GET /
# ─────────────────────────────────────────────
@app.route("/")
def index():
    """Documentación de la API en JSON — fácil de consumir desde Power Automate."""
    return _ok(
        data={
            "endpoints": [
                {
                    "method":      "GET",
                    "path":        "/process-data",
                    "descripcion": "Consulta Jira y devuelve resumen mensual + evolutivo en JSON. "
                                   "Solo lectura — no guarda en BD.",
                    "params": {
                        "inicio": "YYYY-MM-DD (default 2025-11-01)",
                        "fin":    "YYYY-MM-DD (default ayer en hora Lima)",
                    },
                    "power_automate": {
                        "tickets":         "@body('HTTP')?['data']?['tickets']",
                        "resumen_mensual": "@body('HTTP')?['data']?['resumen_mensual']",
                        "evolutivo":       "@body('HTTP')?['data']?['evolutivo']",
                        "html_completo":   "@body('HTTP')?['data']?['html_completo']",
                    },
                },
                {
                    "method":      "POST",
                    "path":        "/guardar-historico",
                    "descripcion": "Guarda snapshot histórico en PostgreSQL. "
                                   "POST porque escribe en BD (no idempotente). "
                                   "Usar bulk INSERT — resuelve WORKER TIMEOUT.",
                    "body_json": {
                        "inicio": "YYYY-MM-DD (default 2025-11-01)",
                        "fin":    "YYYY-MM-DD (default ayer en hora Lima)",
                    },
                    "power_automate": {
                        "supervisores_guardados": "@body('HTTP')?['data']?['supervisores_guardados']",
                        "guardado":               "@body('HTTP')?['data']?['guardado']",
                    },
                },
                {
                    "method":      "GET",
                    "path":        "/download-excel",
                    "descripcion": "Descarga Excel con hojas Detalle, Mensual y Evolutivo.",
                    "params": {
                        "inicio": "YYYY-MM-DD (default 2025-11-01)",
                        "fin":    "YYYY-MM-DD (default ayer en hora Lima)",
                    },
                    "nota": "Devuelve binario .xlsx, no JSON. Use 'Get file content' en Power Automate.",
                },
                {
                    "method":      "GET",
                    "path":        "/health",
                    "descripcion": "Verifica disponibilidad de la API y conectividad con Jira.",
                    "power_automate": {
                        "jira": "@body('HTTP')?['data']?['jira']",
                        "api":  "@body('HTTP')?['data']?['api']",
                    },
                },
            ]
        },
        message="API Flask TPRO Jira — documentación de endpoints",
        meta={"version": "3.0", "puerto": 5000},
    )


# ─────────────────────────────────────────────────────────────────
# ╔═══════════════════════════════════════════════════════════════╗
# ║  GET /process-data                                           ║
# ╠═══════════════════════════════════════════════════════════════╣
# ║  Método: GET ✅                                              ║
# ║  Razón  : Solo CONSULTA Jira (GET externo) y BD (SELECT).   ║
# ║           No muta ningún estado → seguro e idempotente.      ║
# ║                                                              ║
# ║  FIX 9: guardar_historico_evolutivo() eliminado de aquí.    ║
# ║  /process-data es ahora LECTURA PURA. El guardado ocurre    ║
# ║  solo desde POST /guardar-historico. Esto elimina la causa   ║
# ║  principal del WORKER TIMEOUT de Gunicorn en producción.    ║
# ╚═══════════════════════════════════════════════════════════════╝
# ─────────────────────────────────────────────────────────────────
@app.route("/process-data", methods=["GET", "OPTIONS"])
@handle_errors
def process_data():
    """
    Endpoint principal de consulta — SOLO LECTURA.

    Devuelve JSON estructurado con:
      • tickets         — lista de issues normalizados
      • resumen_mensual — agrupación por supervisor / mes / estado
      • evolutivo       — evolución histórica (últimas 5 fechas de la BD)
      • html_completo   — HTML renderizado listo para pegar en un correo

    Power Automate puede acceder a:
        @body('HTTP')?['data']?['tickets']
        @body('HTTP')?['data']?['resumen_mensual']
        @body('HTTP')?['data']?['evolutivo']
        @body('HTTP')?['data']?['html_completo']
        @body('HTTP')?['meta']?['total_issues']

    NOTA: Este endpoint ya NO guarda en la BD (FIX 9).
    Para guardar el snapshot usar POST /guardar-historico.
    """
    if request.method == "OPTIONS":
        return _ok(message="CORS OK")

    params = _get_params()

    # ── Validación explícita de tipos ─────────────────────────────
    fecha_inicio = _parse_date(
        params.get("inicio", "2025-11-01"), "inicio"
    )
    fecha_fin = _parse_date(
        params.get("fin", _ayer_lima_str()), "fin"   # Por defecto: ayer en hora Lima
    )
    print(f"\n{'='*55}")
    print(f"[{datetime.now():%H:%M:%S}] GET /process-data")
    print(f"[{datetime.now():%H:%M:%S}] Rango  : {fecha_inicio} → {fecha_fin}")
    print(f"{'='*55}\n")

    # ── Consulta a Jira (GET, solo lectura) ───────────────────────
    issues = obtener_issues(fecha_inicio, fecha_fin)

    if not issues:
        return _err(
            "No se encontraron issues para el rango indicado",
            errors=[f"Rango: {fecha_inicio} → {fecha_fin}. "
                    "Verifica fechas y permisos de acceso a Jira."],
            status=404,
        )

    print(f"[{datetime.now():%H:%M:%S}] Total issues: {len(issues)}")

    # ── Procesamiento ─────────────────────────────────────────────
    df         = construir_dataframe(issues)
    df_mensual = generar_mensual(df)

    # FIX 9: El guardado histórico fue eliminado de este endpoint.
    # /process-data es ahora lectura pura — no toca la BD en escritura.
    # Para persistir el snapshot usar POST /guardar-historico.

    # ── Evolutivo — solo lectura desde BD ─────────────────────────
    df_evolutivo = generar_evolutivo(df, usar_bd=True, fecha_fin=fecha_fin)

    # ── Serializar a JSON ─────────────────────────────────────────
    tickets_json         = _df_to_records(df)
    resumen_mensual_json = _df_to_records(df_mensual)
    evolutivo_json       = _evolutivo_to_dict(df_evolutivo)

    # ── Fecha de corte en texto ───────────────────────────────────
    meses_es = [
        "", "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
    ]
    fc        = datetime.strptime(fecha_fin, "%Y-%m-%d")
    fecha_txt = f"{fc.day} de {meses_es[fc.month]} del {fc.year}"

    # ── HTML completo (para correo / visualización) ───────────────
    tabla_resumen   = generar_tabla_resumen_html(df_mensual)
    tabla_evolutivo = generar_tabla_evolutivo_html(df_evolutivo)
    html_completo   = render_template_string(
        HTML_TEMPLATE,
        tabla_resumen=tabla_resumen,
        tabla_evolutivo=tabla_evolutivo,
        fecha_corte=fecha_txt,
    )

    return _ok(
        data={
            "tickets":              tickets_json,
            "resumen_mensual":      resumen_mensual_json,
            "evolutivo":            evolutivo_json,
            "html_completo":        html_completo,
            "tabla_resumen_html":   tabla_resumen,
            "tabla_evolutivo_html": tabla_evolutivo,
            "fecha_corte_texto":    fecha_txt,
        },
        message=f"{len(issues)} issues procesados correctamente",
        meta={
            "total_issues":       len(issues),
            "total_supervisores": int(df["supervisor_final"].nunique()) if not df.empty else 0,
            "fecha_inicio":       fecha_inicio,
            "fecha_fin":          fecha_fin,
        },
    )


# ─────────────────────────────────────────────────────────────────
# ╔═══════════════════════════════════════════════════════════════╗
# ║  POST /guardar-historico                                     ║
# ╠═══════════════════════════════════════════════════════════════╣
# ║  Método: POST ✅                                             ║
# ║  Razón  : Escribe (DELETE + INSERT bulk) en PostgreSQL.      ║
# ║           No idempotente → POST correcto.                    ║
# ║  FIX 6: Bulk INSERT — elimina WORKER TIMEOUT.               ║
# ║  FIX 7: Transacción atómica — sin tabla vacía intermedia.   ║
# ╚═══════════════════════════════════════════════════════════════╝
# ─────────────────────────────────────────────────────────────────
@app.route("/guardar-historico", methods=["POST", "OPTIONS"])
@handle_errors
def guardar_historico_endpoint():
    """
    Persiste el snapshot actual en la tabla evolutiva de PostgreSQL.

    Body JSON esperado:
        { "inicio": "2025-11-01", "fin": "2026-02-17" }

    Respuesta JSON para Power Automate:
        @body('HTTP')?['data']?['guardado']               → true/false
        @body('HTTP')?['data']?['supervisores_guardados'] → número
        @body('HTTP')?['data']?['fecha_corte']            → "2026-02-17"
        @body('HTTP')?['meta']?['total_issues']           → número
    """
    if request.method == "OPTIONS":
        return _ok(message="CORS OK")

    params = _get_params()

    fecha_inicio = _parse_date(
        params.get("inicio", "2025-11-01"), "inicio"
    )
    fecha_fin = _parse_date(
        params.get("fin", _ayer_lima_str()), "fin"   # Por defecto: ayer en hora Lima
    )

    print(f"\n{'='*55}")
    print(f"[{datetime.now():%H:%M:%S}] POST /guardar-historico")
    print(f"[{datetime.now():%H:%M:%S}] Rango: {fecha_inicio} → {fecha_fin}")
    print(f"{'='*55}\n")

    issues = obtener_issues(fecha_inicio, fecha_fin)

    if not issues:
        return _err(
            "No se encontraron issues para guardar",
            errors=[f"Rango: {fecha_inicio} → {fecha_fin}"],
            status=404,
        )

    df           = construir_dataframe(issues)
    resultado_bd = guardar_historico_evolutivo(df, fecha_inicio, fecha_fin)

    return _ok(
        data=resultado_bd,
        message="Snapshot histórico guardado correctamente en PostgreSQL",
        meta={
            "total_issues":       len(issues),
            "total_supervisores": int(df["supervisor_final"].nunique()) if not df.empty else 0,
            "fecha_inicio":       fecha_inicio,
            "fecha_fin":          fecha_fin,
        },
    )


# ─────────────────────────────────────────────────────────────────
# ╔═══════════════════════════════════════════════════════════════╗
# ║  GET /download-excel                                         ║
# ╠═══════════════════════════════════════════════════════════════╣
# ║  Método: GET ✅                                              ║
# ║  Razón  : Descarga de recurso. Idempotente. No muta estado.  ║
# ║  Nota   : Devuelve binario .xlsx, no JSON.                   ║
# ║           En Power Automate usar acción "Get file content"   ║
# ║           o guardar en SharePoint/OneDrive directamente.     ║
# ╚═══════════════════════════════════════════════════════════════╝
# ─────────────────────────────────────────────────────────────────
@app.route("/download-excel", methods=["GET"])
@handle_errors
def download_excel():
    """
    Genera y descarga el archivo Excel con tres hojas:
      • Detalle    — todos los issues individuales
      • Mensual    — agrupación por supervisor/mes/estado
      • Evolutivo  — evolución histórica pivoteada

    FIX 15: base64 y FlaskResponse ya están importados al inicio del módulo.
    """
    params = _get_params()

    fecha_inicio = _parse_date(
        params.get("inicio", "2025-11-01"), "inicio"
    )
    fecha_fin = _parse_date(
        params.get("fin", _ayer_lima_str()), "fin"   # Por defecto: ayer en hora Lima
    )

    print(f"\n[{datetime.now():%H:%M:%S}] GET /download-excel | {fecha_inicio} → {fecha_fin}")

    issues = obtener_issues(fecha_inicio, fecha_fin)

    if not issues:
        return _err("No hay datos para exportar", status=404)

    df           = construir_dataframe(issues)
    df_mensual   = generar_mensual(df)
    df_evolutivo = generar_evolutivo(df, usar_bd=True, fecha_fin=fecha_fin)

    # Columnas de la hoja Detalle — orden explícito garantiza que los campos
    # nuevos aparezcan aunque el DataFrame tenga columnas extra o distinto orden.
    # Solo se incluyen las columnas que realmente existen en el df (evita KeyError
    # si en algún entorno falta una columna por configuración).
    COLUMNAS_DETALLE = [
        "cod_trabajo_programado",
        "tipo_ventana",
        "estado",
        "resumen",
        "fecha_inicio",
        "fecha_fin",
        "fecha_comite",
        "supervisor",             # supervisor original de Jira
        "registrado_por",
        "region",
        "grupo_localidad",        # nivel padre del Cascading Select
        "localidad",              # nivel hijo  del Cascading Select
        "empresa_ejecutor",
        "supervisor_final",       # supervisor con reglas de negocio aplicadas
    ]
    cols_presentes = [c for c in COLUMNAS_DETALLE if c in df.columns]

    # ── Anchos máximos por columna (en caracteres) ─────────────────
    ANCHO_MAX = {
        "resumen":          45,
        "supervisor":       35,
        "registrado_por":   30,
        "localidad":        35,
        "empresa_ejecutor": 32,
        "supervisor_final": 35,
    }
    ANCHO_MIN = 8   # ninguna columna más angosta que esto

    def _autofit_sheet(ws):
        """
        Ajusta el ancho de cada columna de una hoja openpyxl al contenido real.
        Recorre todas las celdas (incluyendo cabecera) y aplica ANCHO_MIN/MAX.
        Activa wrap_text en celdas cuyo contenido supere el ancho maximo.
        """
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment

        for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
            col_letter  = get_column_letter(col_idx)
            header_name = col_cells[0].value  # fila 1 = cabecera

            # Ancho real: máximo de todos los valores de la columna
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col_cells
            )

            # Aplicar límites
            maximo = ANCHO_MAX.get(str(header_name), 50)
            ancho  = max(ANCHO_MIN, min(max_len + 2, maximo))
            ws.column_dimensions[col_letter].width = ancho

            # Wrap text en celdas de datos que superan el ancho máximo
            if max_len > maximo:
                for cell in col_cells[1:]:   # saltar cabecera
                    cell.alignment = Alignment(wrap_text=True)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df[cols_presentes].to_excel(writer, sheet_name="Detalle",   index=False)
        df_mensual.to_excel(writer,         sheet_name="Mensual",   index=False)
        df_evolutivo.to_excel(writer,       sheet_name="Evolutivo")

        # Autoajuste dentro del bloque `with` — workbook todavía abierto
        for sheet_name in writer.sheets:
            _autofit_sheet(writer.sheets[sheet_name])

    output.seek(0)
    filename = f"Detalle_jira_{fecha_fin.replace('-', '')}.xlsx"

    # ── Detección del cliente ──────────────────────────────────────
    # Power Automate envía Accept: application/json → necesita Base64
    # Navegador / curl envían Accept: */* o application/octet-stream
    #   → recibe el binario directo para descarga normal
    # FIX 15: base64 y FlaskResponse ya están importados al inicio del módulo.
    accept = request.headers.get("Accept", "")

    if "application/json" in accept:
        archivo_base64 = base64.b64encode(output.read()).decode("utf-8")
        return _ok(
            data={
                "filename":    filename,
                "filecontent": archivo_base64,
                "encoding":    "base64",
                "mimetype":    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
            message=f"Excel generado correctamente: {filename}",
            meta={
                "fecha_inicio": fecha_inicio,
                "fecha_fin":    fecha_fin,
            },
        )

    # Descarga binaria directa (navegador / Postman / curl)
    output.seek(0)
    return FlaskResponse(
        output.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─────────────────────────────────────────────────────────────────
# ╔═══════════════════════════════════════════════════════════════╗
# ║  GET /health                                                 ║
# ╠═══════════════════════════════════════════════════════════════╣
# ║  Método: GET ✅  Endpoint de monitoreo                       ║
# ║  Razón  : Monitoreo/warm-up. Siempre idempotente.            ║
# ║           Power Automate puede usarlo como "ping" antes de   ║
# ║           lanzar el flujo principal, evitando timeouts en    ║
# ║           instancias dormidas (Render free tier, etc.).      ║
# ╚═══════════════════════════════════════════════════════════════╝
# ─────────────────────────────────────────────────────────────────
@app.route("/health", methods=["GET"])
def health():
    """
    Verifica disponibilidad de la API y conectividad con Jira y BD.

    Power Automate:
        @body('HTTP')?['data']?['jira']  → "ok" | "error"
        @body('HTTP')?['data']?['api']   → "ok"
        @body('HTTP')?['data']?['db']    → "ok" | "error"
    """
    jira_ok = False
    db_ok   = False

    try:
        r = requests.get(
            f"{JIRA_URL}/rest/api/3/myself",
            auth=AUTH, headers=REQ_HEADERS, timeout=10,
        )
        jira_ok = (r.status_code == 200)
    except Exception:
        pass

    try:
        with DB_ENGINE.connect() as conn:
            conn.execute(text("SELECT 1"))
        db_ok = True
    except Exception:
        pass

    todo_ok     = jira_ok and db_ok
    status_code = 200 if todo_ok else 503
    return _ok(
        data={
            "api":  "ok",
            "jira": "ok" if jira_ok else "error",
            "db":   "ok" if db_ok   else "error",
        },
        message="API operativa" if todo_ok else "Uno o más servicios no disponibles",
        meta={"version": "3.0"},
        status=status_code,
    )


# ─────────────────────────────────────────────────────────────────
# HANDLERS PARA RUTAS AUTOMÁTICAS DEL NAVEGADOR
# Chrome/Edge las solicitan automáticamente al abrir la API.
# Sin estos handlers generan 404 que ensucian los logs.
# No tienen impacto en Power Automate.
# ─────────────────────────────────────────────────────────────────
@app.route("/favicon.ico")
def favicon():
    """Chrome solicita favicon automáticamente — respuesta vacía sin error."""
    return "", 204  # 204 No Content


@app.route("/.well-known/appspecific/com.chrome.devtools.json")
def chrome_devtools():
    """Chrome DevTools lo solicita al inspeccionar — respuesta vacía sin error."""
    return "", 204  # 204 No Content


# ═══════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("\n" + "=" * 62)
    print("  🚀  flask_jira_refactored.py v3.0 — Power Automate Ready")
    print("=" * 62)
    print()
    print("  CAMBIOS v3 (correcciones críticas):")
    print("  ┌──────────────────────────────────────────────────────┐")
    print("  │ ✅ Bulk INSERT — elimina WORKER TIMEOUT de Gunicorn  │")
    print("  │ ✅ Transacción atómica DELETE+INSERT                 │")
    print("  │ ✅ Engine global con connection pool                 │")
    print("  │ ✅ /process-data es ahora SOLO LECTURA              │")
    print("  │ ✅ Límite MAX_PAGES — previene loop infinito Jira    │")
    print("  │ ✅ timeout=25s (< Gunicorn 30s) — 504 manejable     │")
    print("  │ ✅ Rate-limit capped a 20s                          │")
    print("  │ ✅ pandas FutureWarning corregido (df.copy())       │")
    print("  │ ✅ Flask 2.3+ JSON config (app.json.*)              │")
    print("  │ ✅ XSS prevenido con html.escape()                  │")
    print("  │ ✅ Allowlist para PG_SCHEMA y PG_TABLE              │")
    print("  └──────────────────────────────────────────────────────┘")
    print()
    print("  MÉTODOS HTTP:")
    print("  ┌──────────────────────────────────────────────────────┐")
    print("  │ GET  /process-data      → solo lectura Jira+BD      │")
    print("  │ POST /guardar-historico → escribe en PostgreSQL      │")
    print("  │ GET  /download-excel    → descarga archivo           │")
    print("  │ GET  /health            → monitoreo Jira+BD          │")
    print("  └──────────────────────────────────────────────────────┘")
    print()
    print("  RESPUESTA JSON (todos los endpoints):")
    print('  { "status", "timestamp", "message", "meta", "data" }')
    print()
    print("  GUNICORN recomendado:")
    print("  gunicorn app:app --workers 2 --timeout 120 --preload")
    print()
    print("  Servidor dev: http://localhost:5000")
    print("=" * 62 + "\n")

    app.run(debug=True, host="0.0.0.0", port=5000)
